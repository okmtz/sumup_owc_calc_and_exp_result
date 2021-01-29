import glob
import csv
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

def main():
    print('####################################################')
    print('now loading file')
    print('####################################################')

    files = glob.glob("*")
    calc_csv = ''
    exp_data_xlsx = ''
    for file in files:
        if ('.csv' in file):
            calc_csv = file
        elif ('.xlsx' in file):
            exp_data_xlsx = file
        else:
            pass

    print('####################################################')
    print(f'load data {exp_data_xlsx}')
    print('####################################################')
    exp_work_book = load_workbook(exp_data_xlsx, data_only=True)
    exp_sheet_name_re = re.search(r'A\d+-T\d+', exp_data_xlsx)
    exp_sheet_name = exp_sheet_name_re.group()
    exp_sheet = exp_work_book[exp_sheet_name]
    max_row_num = exp_sheet.max_row
    exp_time_list = [(exp_sheet.cell(row=i+12, column=9).value) for i in range(max_row_num)]
    exp_pressure_list = [(exp_sheet.cell(row=i+12, column=11).value) for i in range(max_row_num)]

    print('#####################################################')
    print(f'convert calculated csv file {calc_csv} to excel')
    print('#####################################################')
    save_file = exp_sheet_name + '-sum.xlsx'
    with open(calc_csv) as f:
        load_data = pd.read_csv(f, index_col=0)
        load_data.to_excel(save_file, encoding='utf-8')

    sum_workbook = load_workbook(save_file)
    sum_sheet = sum_workbook['Sheet1']

    print('#####################################################')
    print(f'write experiment data to {save_file}')
    print('#####################################################')

    i = 0
    for time, pres in zip(exp_time_list, exp_pressure_list):
        i += 1
        sum_sheet[f"E{i}"] = time
        sum_sheet[f"F{i}"] = pres

    print('#####################################################')
    print('now drawing graph')
    print('#####################################################')
    chart = ScatterChart()
    max_row_num = sum_sheet.max_row
    calc_y = Reference(sum_workbook['Sheet1'], min_col=4, max_col=4, min_row=2, max_row=max_row_num)
    calc_x = Reference(sum_workbook['Sheet1'], min_col=2, max_col=2, min_row=2, max_row=max_row_num)
    calc_series = Series(calc_y, calc_x, title='calc')
    calc_series.graphicalProperties.line.noFill = True
    calc_series.marker.symbol = "auto"

    exp_y = Reference(sum_workbook['Sheet1'], min_col=6, max_col=6, min_row=2, max_row=len(exp_pressure_list))
    exp_x = Reference(sum_workbook['Sheet1'], min_col=5, max_col=5, min_row=2, max_row=len(exp_time_list))
    exp_series = Series(exp_y, exp_x, title='exp')

    chart.append(calc_series)
    chart.append(exp_series)
    sum_workbook.create_sheet(title='compare_flow')
    sum_workbook['compare_flow'].add_chart(chart, "A6")
    sum_workbook.save(save_file)
    print('#####################################################')
    print(f'sumup file {save_file} finished')
    print('#####################################################')

if __name__=='__main__':
    main()